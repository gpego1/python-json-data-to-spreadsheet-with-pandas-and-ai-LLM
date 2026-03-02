import pandas as pd
import os
from datetime import datetime
import logging
from typing import List, Dict, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self, output_path="/app/output"):
        self.output_path = output_path
        os.makedirs(output_path, exist_ok=True)
        
        self.header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.subheader_fill = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")
        self.alternate_fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
        self.total_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def format_brl(self, value):
        try:
            if pd.isna(value) or value is None:
                return "R$ 0,00"
            return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return "R$ 0,00"
    
    def apply_formatting(self, worksheet, df, sheet_name):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            try:
                max_length = df[col_name].astype(str).map(len).max()
            except:
                max_length = len(str(col_name))
            
            col_width = min(max(max_length, len(str(col_name))) + 2, 45)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = col_width
        
        for row_idx in range(2, worksheet.max_row + 1):
            fill = self.alternate_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = self.border
                if fill:
                    cell.fill = fill
                
                col_name = df.columns[col_idx - 1]
                if any(x in col_name.lower() for x in ['preço', 'total', 'unitário', 'r$', 'valor']):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif any(x in col_name.lower() for x in ['id', 'quantidade']):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        worksheet.freeze_panes = 'A2'
    
    def add_totals_row(self, worksheet, df, sheet_name):
        if sheet_name != 'Itens dos Pedidos':
            return
        
        total_row = worksheet.max_row + 2
        summary_row = total_row + 2
        
        cell = worksheet.cell(row=total_row, column=1, value="RESUMO DOS TOTAIS")
        cell.font = Font(bold=True, size=11)
        cell.fill = self.subheader_fill
        cell.font = Font(color="FFFFFF", bold=True)
        
        def extract_number(value):
            if isinstance(value, str):
                return float(value.replace('R$', '').replace('.', '').replace(',', '.').strip())
            return float(value) if value else 0
        
        total_quantity = df['Quantidade'].sum() if 'Quantidade' in df.columns else 0
        total_item_value = sum(extract_number(v) for v in df['Total do Item (R$)']) if 'Total do Item (R$)' in df.columns else 0
        total_order_value = sum(extract_number(v) for v in df['Total do Pedido (R$)']) if 'Total do Pedido (R$)' in df.columns else 0
        
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == 'Quantidade':
                cell = worksheet.cell(row=total_row + 1, column=col_idx, value=total_quantity)
                cell.font = Font(bold=True)
                cell.number_format = '#,##0'
                cell.fill = self.total_fill
                cell.border = self.border
            elif col_name == 'Total do Item (R$)':
                cell = worksheet.cell(row=total_row + 1, column=col_idx, value=total_item_value)
                cell.font = Font(bold=True)
                cell.number_format = 'R$ #,##0.00'
                cell.fill = self.total_fill
                cell.border = self.border
            elif col_name == 'Total do Pedido (R$)':
                cell = worksheet.cell(row=total_row + 1, column=col_idx, value=total_order_value)
                cell.font = Font(bold=True)
                cell.number_format = 'R$ #,##0.00'
                cell.fill = self.total_fill
                cell.border = self.border
        
        cell = worksheet.cell(row=summary_row, column=1, value="MÉDIAS")
        cell.font = Font(bold=True, size=11)
        cell.fill = self.subheader_fill
        cell.font = Font(color="FFFFFF", bold=True)
        
        avg_item = total_item_value / len(df) if len(df) > 0 else 0
        avg_quantity = total_quantity / len(df) if len(df) > 0 else 0
        
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == 'Quantidade':
                cell = worksheet.cell(row=summary_row + 1, column=col_idx, value=round(avg_quantity, 1))
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.0'
                cell.border = self.border
            elif col_name == 'Total do Item (R$)':
                cell = worksheet.cell(row=summary_row + 1, column=col_idx, value=avg_item)
                cell.font = Font(bold=True)
                cell.number_format = 'R$ #,##0.00'
                cell.border = self.border
    
    def add_summary_statistics(self, writer, df):
        stats_data = []
        
        total_orders = int(df['order_id'].nunique()) if 'order_id' in df.columns else 0
        total_items = int(df['quantity'].sum()) if 'quantity' in df.columns else 0
        total_value = float(df['order_total'].sum()) if 'order_total' in df.columns else 0
        
        stats_data.append(['Total de Pedidos', str(total_orders)])
        stats_data.append(['Total de Itens Vendidos', str(total_items)])
        stats_data.append(['Valor Total das Vendas', self.format_brl(total_value)])
        
        if total_orders > 0:
            stats_data.append(['Ticket Médio por Pedido', self.format_brl(total_value / total_orders)])
        else:
            stats_data.append(['Ticket Médio por Pedido', self.format_brl(0)])
        
        if len(df) > 0:
            stats_data.append(['Valor Médio por Item', self.format_brl(total_value / len(df))])
            stats_data.append(['Quantidade Média por Item', f"{total_items/len(df):.1f}"])
        else:
            stats_data.append(['Valor Médio por Item', self.format_brl(0)])
            stats_data.append(['Quantidade Média por Item', "0"])
        
        if 'order_total' in df.columns and len(df) > 0:
            stats_data.append(['Maior Pedido', self.format_brl(df['order_total'].max())])
            stats_data.append(['Menor Pedido', self.format_brl(df['order_total'].min())])
        else:
            stats_data.append(['Maior Pedido', self.format_brl(0)])
            stats_data.append(['Menor Pedido', self.format_brl(0)])
        
        if 'occurred_at' in df.columns and len(df) > 0:
            dates = pd.to_datetime(df['occurred_at'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
            valid_dates = dates.dropna()
            if len(valid_dates) > 0:
                stats_data.append(['Data Mais Antiga', valid_dates.min().strftime('%d/%m/%Y %H:%M')])
                stats_data.append(['Data Mais Recente', valid_dates.max().strftime('%d/%m/%Y %H:%M')])
            else:
                stats_data.append(['Data Mais Antiga', '-'])
                stats_data.append(['Data Mais Recente', '-'])
        else:
            stats_data.append(['Data Mais Antiga', '-'])
            stats_data.append(['Data Mais Recente', '-'])
        
        stats_df = pd.DataFrame(stats_data, columns=['Métrica', 'Valor'])
        stats_df.to_excel(writer, sheet_name='Estatísticas', index=False)
        
        worksheet = writer.sheets['Estatísticas']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 25
        
        for row in range(1, len(stats_df) + 2):
            for col in [1, 2]:
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.border
                if row == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif col == 2 and row > 1:
                    cell.alignment = Alignment(horizontal='right')
    
    def generate_excel(self, rows: List[Dict], order_id: Optional[int] = None) -> Optional[str]:
        if not rows:
            logger.warning("Nenhuma linha para gerar Excel")
            return None
        
        try:
            df = pd.DataFrame(rows)
            
            required_columns = ['order_id', 'user_id', 'product_id', 'quantity', 'unit_price', 'item_total', 'order_total', 'occurred_at']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 0 if col != 'occurred_at' else ''
            
            df['order_id'] = pd.to_numeric(df['order_id'], errors='coerce').fillna(0).astype(int)
            df['user_id'] = pd.to_numeric(df['user_id'], errors='coerce').fillna(0).astype(int)
            df['product_id'] = pd.to_numeric(df['product_id'], errors='coerce').fillna(0).astype(int)
            df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)
            df['unit_price'] = pd.to_numeric(df['unit_price'], errors='coerce').fillna(0.0).astype(float)
            df['item_total'] = pd.to_numeric(df['item_total'], errors='coerce').fillna(0.0).astype(float)
            df['order_total'] = pd.to_numeric(df['order_total'], errors='coerce').fillna(0.0).astype(float)
            
            try:
                df['occurred_at'] = pd.to_datetime(df['occurred_at'], errors='coerce')
                df['occurred_at'] = df['occurred_at'].dt.strftime('%d/%m/%Y %H:%M:%S')
            except:
                df['occurred_at'] = df['occurred_at'].astype(str)
            
            df_display = pd.DataFrame()
            df_display['ID do Pedido'] = df['order_id']
            df_display['ID do Usuário'] = df['user_id']
            df_display['ID do Produto'] = df['product_id']
            df_display['Quantidade'] = df['quantity']
            df_display['Preço Unitário (R$)'] = df['unit_price'].apply(self.format_brl)
            df_display['Total do Item (R$)'] = df['item_total'].apply(self.format_brl)
            df_display['Total do Pedido (R$)'] = df['order_total'].apply(self.format_brl)
            df_display['Data do Pedido'] = df['occurred_at']
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_pedidos_{timestamp}.xlsx" if not order_id else f"pedido_{order_id}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_display.to_excel(writer, sheet_name='Itens dos Pedidos', index=False)
                self.apply_formatting(writer.sheets['Itens dos Pedidos'], df_display, 'Itens dos Pedidos')
                self.add_totals_row(writer.sheets['Itens dos Pedidos'], df_display, 'Itens dos Pedidos')
                
                if not df.empty and 'order_id' in df.columns:
                    summary_df = df.groupby('order_id').agg({
                        'user_id': 'first',
                        'order_total': 'first',
                        'item_total': 'sum',
                        'order_id': 'count',
                        'occurred_at': 'first'
                    }).reset_index()
                    
                    summary_display = pd.DataFrame()
                    summary_display['ID do Pedido'] = summary_df['order_id']
                    summary_display['ID do Usuário'] = summary_df['user_id']
                    summary_display['Quantidade de Itens'] = summary_df['order_id']
                    summary_display['Total do Pedido (R$)'] = summary_df['order_total'].apply(self.format_brl)
                    summary_display['Soma dos Itens (R$)'] = summary_df['item_total'].apply(self.format_brl)
                    
                    try:
                        dates = pd.to_datetime(summary_df['occurred_at'], errors='coerce')
                        summary_display['Data do Pedido'] = dates.dt.strftime('%d/%m/%Y %H:%M:%S')
                    except:
                        summary_display['Data do Pedido'] = summary_df['occurred_at'].astype(str)
                    
                    summary_display.to_excel(writer, sheet_name='Resumo por Pedido', index=False)
                    self.apply_formatting(writer.sheets['Resumo por Pedido'], summary_display, 'Resumo por Pedido')
                
                if not df.empty and 'product_id' in df.columns:
                    product_df = df.groupby('product_id').agg({
                        'quantity': 'sum',
                        'item_total': 'sum',
                        'order_id': 'nunique'
                    }).reset_index()
                    
                    product_display = pd.DataFrame()
                    product_display['ID do Produto'] = product_df['product_id']
                    product_display['Quantidade Total'] = product_df['quantity']
                    product_display['Valor Total (R$)'] = product_df['item_total'].apply(self.format_brl)
                    product_display['Pedidos Únicos'] = product_df['order_id']
                    
                    product_display.to_excel(writer, sheet_name='Resumo por Produto', index=False)
                    self.apply_formatting(writer.sheets['Resumo por Produto'], product_display, 'Resumo por Produto')
                
                self.add_summary_statistics(writer, df)
            
            logger.info(f"Excel gerado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {e}")
            import traceback
            traceback.print_exc()
            return None